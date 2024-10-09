import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging
from tkinter import Tk, filedialog, messagebox, Toplevel, Button, Frame, Scrollbar, Checkbutton, IntVar, Canvas, Label, \
    Entry, ttk
import random

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def sanitize_filename(filename):
    return ''.join(c for c in filename if c.isalnum() or c in (' ', '_', '-')).rstrip()


def extract_roll_numbers_from_file(file_path):
    try:
        student_data = pd.read_excel(file_path)
        roll_numbers = student_data['Enrollment Number'].tolist()
        logging.info(f"Extracted {len(roll_numbers)} roll numbers from {file_path}")
        return roll_numbers
    except Exception as e:
        logging.error(f"Error extracting roll numbers from {file_path}: {str(e)}")
        return []


def generate_combined_roll_numbers(roll_numbers_ii, roll_numbers_iii):
    return [(roll_numbers_ii[i] if i < len(roll_numbers_ii) else '',
             roll_numbers_iii[i] if i < len(roll_numbers_iii) else '')
            for i in range(max(len(roll_numbers_ii), len(roll_numbers_iii)))]


def generate_seating_excel(output_file, roll_numbers, exam_codes, exam_date, exam_time, selected_faculty, room_layouts,
                           is_fourth_year=False):
    wb = Workbook()
    ws = wb.active

    # Add header information
    header_info = [
        f'Exam Codes: {", ".join(exam_codes)}',
        f'Date: {exam_date}',
        f'Time: {exam_time}',
        f'Room: {"4th Year" if is_fourth_year else "Combined 2nd and 3rd Year"}'
    ]

    for i, line in enumerate(header_info):
        cell = ws.cell(row=i + 1, column=1, value=line)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    current_row = len(header_info) + 2  # Start after header
    roll_index = 0
    used_faculty = set()

    for room_name, (total_benches, rows, benches_per_row) in room_layouts.items():
        ws.cell(row=current_row, column=1, value=f"{room_name} Seating Arrangement")
        current_row += 1

        # Select two different faculty members for the room
        available_faculty = [f for f in selected_faculty if f not in used_faculty]
        if len(available_faculty) < 2:
            available_faculty = selected_faculty  # Reset if we've used all faculty
        selected_room_faculty = random.sample(available_faculty, min(2, len(available_faculty)))
        used_faculty.update(selected_room_faculty)

        ws.cell(row=current_row, column=1, value="Selected Faculty:")
        current_row += 1
        for faculty in selected_room_faculty:
            ws.cell(row=current_row, column=1, value=faculty)
            current_row += 1

        # Add column headers (Row numbers)
        for row in range(1, rows + 1):
            ws.cell(row=current_row, column=row + 1, value=f"Row {row}")

        current_row += 1

        # Add row headers (Bench numbers) and fill in student data
        for bench in range(1, benches_per_row + 1):
            ws.cell(row=current_row, column=1, value=f"Bench {bench}")
            for row in range(1, rows + 1):
                if roll_index < len(roll_numbers):
                    cell = ws.cell(row=current_row, column=row + 1)
                    student = roll_numbers[roll_index]
                    cell.value = f"{student[0]}\n{student[1]}"
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'),
                                         bottom=Side(style='thin'), right=Side(style='thin'))
                    roll_index += 1
            current_row += 1

        current_row += 2  # Add space between rooms

    # Auto-fit columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    logging.info(f'Saved: {output_file}')


def extract_faculty_data(faculty_file):
    try:
        faculty_data = pd.read_excel(faculty_file)

        if 'FacultyID' not in faculty_data.columns:
            raise ValueError("The faculty file must contain a 'FacultyID' column.")

        if 'Name' not in faculty_data.columns:
            raise ValueError("The faculty file must contain a 'Name' column.")

        return faculty_data[['FacultyID', 'Name']].dropna()
    except Exception as e:
        logging.error(f"Error extracting faculty data from {faculty_file}: {str(e)}")
        return pd.DataFrame(columns=['FacultyID', 'Name'])


def load_room_data(file_path):
    try:
        room_data = pd.read_excel(file_path)
        room_layouts = {}
        for _, row in room_data.iterrows():
            room_layouts[row['Room Name']] = (int(row['Total Benches']), int(row['Rows']), int(row['Benches per Row']))
        return room_layouts
    except Exception as e:
        logging.error(f"Error loading room data from {file_path}: {str(e)}")
        return {}


def generate_exam_files(timetable_file, student_files, faculty_file, room_file, output_dir, selected_faculty):
    try:
        timetable = pd.read_excel(timetable_file)
        logging.info(f"Loaded timetable from {timetable_file}")

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.info(f"Created output directory: {output_dir}")

        roll_numbers_ii = extract_roll_numbers_from_file(student_files.get('2nd', ''))
        roll_numbers_iii = extract_roll_numbers_from_file(student_files.get('3rd', ''))
        roll_numbers_iv = extract_roll_numbers_from_file(student_files.get('4th', ''))

        room_layouts = load_room_data(room_file)
        if not room_layouts:
            messagebox.showerror("Error", "Failed to load room data. Please check the room file format.")
            return

        files_generated = 0
        total_capacity = sum(layout[0] for layout in room_layouts.values())

        for _, row in timetable.iterrows():
            exam_date = row['Date'].strftime('%Y-%m-%d') if isinstance(row['Date'], pd.Timestamp) else row['Date']
            exam_time = row['Time'].strftime('%H-%M-%S') if isinstance(row['Time'], pd.Timestamp) else row['Time']
            exam_ii_year = row['B.Tech II Year']
            exam_iii_year = row['B.Tech III Year']
            exam_iv_year = row['B.Tech IV Year']

            if pd.notna(exam_ii_year) and pd.notna(exam_iii_year):
                combined_roll_numbers = generate_combined_roll_numbers(roll_numbers_ii, roll_numbers_iii)
                exam_codes = [exam_ii_year, exam_iii_year]
                output_file = f'{output_dir}/{sanitize_filename("_".join(exam_codes))}_{sanitize_filename(exam_date)}_{sanitize_filename(exam_time)}.xlsx'

                allocated_students = combined_roll_numbers[:total_capacity]
                overflow_students = combined_roll_numbers[total_capacity:]

                generate_seating_excel(output_file, allocated_students, exam_codes, exam_date, exam_time,
                                       selected_faculty, room_layouts)
                files_generated += 1

                if overflow_students:
                    logging.warning(
                        f"Overflow of {len(overflow_students)} students for exam {', '.join(exam_codes)} on {exam_date} at {exam_time}")
                    messagebox.showwarning("Overflow Warning",
                                           f"There are {len(overflow_students)} students without seats for exam {', '.join(exam_codes)} on {exam_date} at {exam_time}")

            elif pd.notna(exam_ii_year) and roll_numbers_ii:
                output_file = f'{output_dir}/{sanitize_filename(exam_ii_year)}_{sanitize_filename(exam_date)}_{sanitize_filename(exam_time)}.xlsx'

                allocated_students = [(roll, '') for roll in roll_numbers_ii[:total_capacity]]
                overflow_students = roll_numbers_ii[total_capacity:]

                generate_seating_excel(output_file, allocated_students, [exam_ii_year], exam_date, exam_time,
                                       selected_faculty, room_layouts)
                files_generated += 1

                if overflow_students:
                    logging.warning(
                        f"Overflow of {len(overflow_students)} students for exam {exam_ii_year} on {exam_date} at {exam_time}")
                    messagebox.showwarning("Overflow Warning",
                                           f"There are {len(overflow_students)} students without seats for exam {exam_ii_year} on {exam_date} at {exam_time}")

            elif pd.notna(exam_iii_year) and roll_numbers_iii:
                output_file = f'{output_dir}/{sanitize_filename(exam_iii_year)}_{sanitize_filename(exam_date)}_{sanitize_filename(exam_time)}.xlsx'

                allocated_students = [(roll, '') for roll in roll_numbers_iii[:total_capacity]]
                overflow_students = roll_numbers_iii[total_capacity:]

                generate_seating_excel(output_file, allocated_students, [exam_iii_year], exam_date, exam_time,
                                       selected_faculty, room_layouts)
                files_generated += 1

                if overflow_students:
                    logging.warning(
                        f"Overflow of {len(overflow_students)} students for exam {exam_iii_year} on {exam_date} at {exam_time}")
                    messagebox.showwarning("Overflow Warning",
                                           f"There are {len(overflow_students)} students without seats for exam {exam_iii_year} on {exam_date} at {exam_time}")

            elif pd.notna(exam_iv_year) and roll_numbers_iv:
                output_file = f'{output_dir}/{sanitize_filename(exam_iv_year)}_{sanitize_filename(exam_date)}_{sanitize_filename(exam_time)}.xlsx'

                allocated_students = [(roll, '') for roll in roll_numbers_iv[:total_capacity]]
                overflow_students = roll_numbers_iv[total_capacity:]

                generate_seating_excel(output_file, allocated_students, [exam_iv_year], exam_date, exam_time,
                                       selected_faculty, room_layouts, is_fourth_year=True)
                files_generated += 1

                if overflow_students:
                    logging.warning(
                        f"Overflow of {len(overflow_students)} students for exam {exam_iv_year} on {exam_date} at {exam_time}")
                    messagebox.showwarning("Overflow Warning",
                                           f"There are {len(overflow_students)} students without seats for exam {exam_iv_year} on {exam_date} at {exam_time}")

        logging.info(f"Generated {files_generated} files in total")
        messagebox.showinfo("Process Completed", f"Generated {files_generated} files in {output_dir}")
    except Exception as e:
        logging.error(f"Error in generate_exam_files: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


class ExamSeatingGUI:
    def __init__(self, master):
        self.master = master
        master.title("Exam Seating Arrangement Generator")
        master.geometry("800x600")

        self.notebook = ttk.Notebook(master)
        self.notebook.pack(expand=True, fill='both', padx=10, pady=10)

        self.instructions_frame = ttk.Frame(self.notebook)
        self.upload_frame = ttk.Frame(self.notebook)

        self.notebook.add(self.instructions_frame, text='Instructions')
        self.notebook.add(self.upload_frame, text='Upload Files')

        self.create_instructions_tab()
        self.create_upload_tab()

    def create_instructions_tab(self):
        instructions = [
            ("Timetable", "Date, Time, B.Tech II Year, B.Tech III Year, B.Tech IV Year"),
            ("Student Files", "Enrollment Number"),
            ("Faculty", "FacultyID, Name"),
            ("Rooms", "Room Name, Total Benches, Rows, Benches per Row")
        ]

        tree = ttk.Treeview(self.instructions_frame, columns=('File Type', 'Required Columns'), show='headings')
        tree.heading('File Type', text='File Type')
        tree.heading('Required Columns', text='Required Columns')

        for item in instructions:
            tree.insert('', 'end', values=item)

        tree.pack(expand=True, fill='both', padx=10, pady=10)

    def create_upload_tab(self):
        self.files = {
            'timetable': None,
            'students2nd': None,
            'students3rd': None,
            'students4th': None,
            'faculty': None,
            'rooms': None
        }

        for i, (key, _) in enumerate(self.files.items()):
            frame = ttk.Frame(self.upload_frame)
            frame.grid(row=i, column=0, padx=10, pady=5, sticky='w')

            label = ttk.Label(frame, text=f"{key.capitalize()} File:")
            label.grid(row=0, column=0, padx=5, pady=5, sticky='w')

            button = ttk.Button(frame, text="Select File", command=lambda k=key: self.select_file(k))
            button.grid(row=0, column=1, padx=5, pady=5)

            self.files[key] = ttk.Label(frame, text="No file selected")

            self.files[key].grid(row=0, column=2, padx=5, pady=5, sticky='w')

        output_frame = ttk.Frame(self.upload_frame)
        output_frame.grid(row=len(self.files), column=0, padx=10, pady=5, sticky='w')

        output_label = ttk.Label(output_frame, text="Output Directory:")
        output_label.grid(row=0, column=0, padx=5, pady=5, sticky='w')

        self.output_entry = ttk.Entry(output_frame, width=50)
        self.output_entry.grid(row=0, column=1, padx=5, pady=5)

        output_button = ttk.Button(output_frame, text="Select Directory", command=self.select_output_dir)
        output_button.grid(row=0, column=2, padx=5, pady=5)

        generate_button = ttk.Button(self.upload_frame, text="Generate Seating Arrangements",
                                     command=self.generate_arrangements)
        generate_button.grid(row=len(self.files) + 1, column=0, padx=10, pady=20)

    def select_file(self, file_type):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.files[file_type].config(text=file_path)

    def select_output_dir(self):
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.output_entry.delete(0, 'end')
            self.output_entry.insert(0, dir_path)

    def generate_arrangements(self):
        files = {k: v.cget("text") for k, v in self.files.items() if v.cget("text") != "No file selected"}
        output_dir = self.output_entry.get()

        if not all(files.values()) or not output_dir:
            messagebox.showerror("Error", "Please select all required files and specify an output directory.")
            return

        faculty_data = extract_faculty_data(files['faculty'])
        selected_faculty = self.select_faculty_members(faculty_data)

        if selected_faculty:
            generate_exam_files(
                files['timetable'],
                {'2nd': files['students2nd'], '3rd': files['students3rd'], '4th': files['students4th']},
                files['faculty'],
                files['rooms'],
                output_dir,
                selected_faculty
            )

    def select_faculty_members(self, faculty_data):
        top = Toplevel(self.master)
        top.title("Select Faculty Members")

        frame = Frame(top)
        frame.pack(fill="both", expand=True)

        canvas = Canvas(frame)
        scrollbar = Scrollbar(frame, orient="vertical", command=canvas.yview)
        scrollable_frame = Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        faculty_vars = {}
        for _, faculty in faculty_data.iterrows():
            var = IntVar()
            faculty_tuple = (faculty['FacultyID'], faculty['Name'])
            faculty_vars[faculty_tuple] = var
            cb = Checkbutton(scrollable_frame,
                             text=f"{faculty['FacultyID']} - {faculty['Name']}",
                             variable=var)
            cb.pack(anchor="w")

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        selected_faculty = []

        def on_ok():
            nonlocal selected_faculty
            selected_faculty = [f"{faculty[0]} - {faculty[1]}" for faculty, var in faculty_vars.items() if var.get()]
            top.destroy()

        ok_button = Button(top, text="OK", command=on_ok)
        ok_button.pack()

        top.wait_window(top)
        return selected_faculty


def main():
    root = Tk()
    gui = ExamSeatingGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
